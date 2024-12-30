import os
import requests
from dotenv import load_dotenv
from datetime import datetime
import time
import threading
import openpyxl
import sys

# ====== ЗАГРУЗКА НАСТРОЕК ====== #
load_dotenv()

FACEBOOK_PAGE_ACCESS_TOKEN = os.getenv("FACEBOOK_PAGE_ACCESS_TOKEN")
FACEBOOK_PAGE_ID = os.getenv("FACEBOOK_PAGE_ID")
PROXY_URL = os.getenv("PROXY_URL")  # Если нужен прокси

# Настраиваем прокси
proxy_dict = None
if PROXY_URL:
    proxy_dict = {
        "http": PROXY_URL,
        "https": PROXY_URL
    }

# ====== ФАЙЛЫ, ГДЕ ХРАНИМ ДАННЫЕ ====== #
PROCESSED_COMMENTS_FILE = "processed_comments.txt"  # Уже обработанные комментарии
POSTED_POSTS_FILE = "posted_posts.txt"              # Наши собственные посты
STATS_XLSX_FILE = "post_stats.xlsx"                 # Статистика (лайки, комментарии, шеры, контент)

# ====== ПАРАМЕТРЫ ПОВЕДЕНИЯ БОТА ====== #
CHECK_INTERVAL = 30  # каждые 30 секунд обновляем статистику и обрабатываем новые комментарии
REPLY_TEMPLATE = """Привет, {author}!
Спасибо за комментарий, {comment}
Мы ценим вашу активность!
"""

def log_action(action: str, info: str):
    """
    Пишем действия в log_facebook.txt с датой/временем.
    """
    with open("log_facebook.txt", "a", encoding="utf-8") as f:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        f.write(f"[{now}] {action}: {info}\n")


def load_processed_comments() -> set:
    """
    Загружаем ID уже обработанных комментариев из файла.
    """
    s = set()
    if os.path.exists(PROCESSED_COMMENTS_FILE):
        with open(PROCESSED_COMMENTS_FILE, "r", encoding="utf-8") as f:
            for line in f:
                c_id = line.strip()
                if c_id:
                    s.add(c_id)
    return s


def save_processed_comment(comment_id: str):
    """
    Дописываем ID комментария в файл, чтобы не обрабатывать повторно.
    """
    with open(PROCESSED_COMMENTS_FILE, "a", encoding="utf-8") as f:
        f.write(comment_id + "\n")


def load_posted_posts() -> list:
    """
    Загружаем из файла все посты, которые мы опубликовали.
    Формат в файле: <post_id> (одна строка — один пост).
    """
    posts = []
    if os.path.exists(POSTED_POSTS_FILE):
        with open(POSTED_POSTS_FILE, "r", encoding="utf-8") as f:
            for line in f:
                p_id = line.strip()
                if p_id:
                    posts.append(p_id)
    return posts


def save_posted_post(post_id: str):
    """
    Сохраняем ID нового поста в файл, если его там ещё нет,
    чтобы не было дублирования.
    """
    posts = load_posted_posts()
    if post_id not in posts:
        with open(POSTED_POSTS_FILE, "a", encoding="utf-8") as f:
            f.write(post_id + "\n")


def init_stats_xlsx():
    """
    Проверяем, существует ли STATS_XLSX_FILE.
    Если нет — создаём новую книгу с листом 'Stats' и шапкой столбцов:
    A - post_id, B - content, C - likes, D - comments, E - shares, F - last_update
    """
    if not os.path.exists(STATS_XLSX_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stats"
        # Шапка
        ws["A1"] = "post_id"
        ws["B1"] = "content"        # <-- столбец для текста поста
        ws["C1"] = "likes"
        ws["D1"] = "comments"
        ws["E1"] = "shares"
        ws["F1"] = "last_update"
        wb.save(STATS_XLSX_FILE)


def update_post_stats(post_id: str, likes: int, comments: int, shares: int, content: str = None):
    """
    Обновляет (или создаёт) строку для поста в Excel,
    где:
      A - post_id
      B - content (заполняется, если новый пост и content не пуст)
      C - likes
      D - comments
      E - shares
      F - last_update
    """
    wb = openpyxl.load_workbook(STATS_XLSX_FILE)
    ws = wb["Stats"]

    # Ищем существующую строку (где A=post_id)
    found_row = None
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value == post_id:
            found_row = row
            break

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if found_row is None:
        # Создаём новую
        found_row = ws.max_row + 1
        ws.cell(row=found_row, column=1).value = post_id
        # Если при первом создании передан контент, запишем его
        if content:
            ws.cell(row=found_row, column=2).value = content

    # Обновим (likes, comments, shares, last_update)
    ws.cell(row=found_row, column=3).value = likes
    ws.cell(row=found_row, column=4).value = comments
    ws.cell(row=found_row, column=5).value = shares
    ws.cell(row=found_row, column=6).value = now_str

    wb.save(STATS_XLSX_FILE)


# =============================================================================
#                     РАБОТА С FACEBOOK (API)
# =============================================================================

def post_to_facebook(message: str) -> str:
    """
    Публикует новый пост на страницу, возвращает ID или сообщение об ошибке.
    Также сразу добавляем запись в Excel и posted_posts.txt
    """
    if not FACEBOOK_PAGE_ACCESS_TOKEN or not FACEBOOK_PAGE_ID:
        error_msg = "Нет токена или ID страницы!"
        log_action("post_to_facebook", error_msg)
        return error_msg

    url = f"https://graph.facebook.com/{FACEBOOK_PAGE_ID}/feed"
    params = {"access_token": FACEBOOK_PAGE_ACCESS_TOKEN}
    data = {
        "message": message,
        "published": "true"
    }
    try:
        r = requests.post(url, params=params, data=data, proxies=proxy_dict, timeout=30)
        if r.status_code == 200:
            new_post_id = r.json().get("id", "")
            msg = f"Пост опубликован! ID: {new_post_id}"
            log_action("post_to_facebook", msg)

            # Сохраним пост в posted_posts.txt (без дублирования)
            save_posted_post(new_post_id)

            # Инициализируем статистику (0,0,0) + content=message в Excel
            update_post_stats(
                post_id=new_post_id,
                likes=0,
                comments=0,
                shares=0,
                content=message
            )
            return new_post_id
        else:
            err = f"Ошибка постинга: {r.status_code} {r.text}"
            log_action("post_to_facebook", err)
            return err
    except Exception as e:
        err = f"Исключение: {e}"
        log_action("post_to_facebook", err)
        return str(err)


def get_post_insights(post_id: str):
    """
    Возвращает (likes, comments, shares) для указанного поста.
    Для likes/comments используем поля summary=1 (edge), для shares - отдельное поле.
    """
    likes_count = 0
    comments_count = 0
    shares_count = 0

    if not FACEBOOK_PAGE_ACCESS_TOKEN:
        return 0, 0, 0

    # Likes (/post_id/likes?summary=true)
    url_likes = f"https://graph.facebook.com/{post_id}/likes"
    params_likes = {
        "summary": "true",
        "access_token": FACEBOOK_PAGE_ACCESS_TOKEN
    }
    try:
        r_likes = requests.get(url_likes, params=params_likes, proxies=proxy_dict, timeout=30)
        if r_likes.status_code == 200:
            data_likes = r_likes.json()
            summary_likes = data_likes.get("summary", {})
            likes_count = summary_likes.get("total_count", 0)
        else:
            log_action("get_post_insights", f"Ошибка likes {post_id}: {r_likes.status_code} {r_likes.text}")
    except Exception as e:
        log_action("get_post_insights", f"Исключение likes {post_id}: {e}")

    # Comments (/post_id/comments?summary=true)
    url_comments = f"https://graph.facebook.com/{post_id}/comments"
    params_comments = {
        "summary": "true",
        "access_token": FACEBOOK_PAGE_ACCESS_TOKEN
    }
    try:
        r_comments = requests.get(url_comments, params=params_comments, proxies=proxy_dict, timeout=30)
        if r_comments.status_code == 200:
            data_com = r_comments.json()
            summary_com = data_com.get("summary", {})
            comments_count = summary_com.get("total_count", 0)
        else:
            log_action("get_post_insights", f"Ошибка comments {post_id}: {r_comments.status_code} {r_comments.text}")
    except Exception as e:
        log_action("get_post_insights", f"Исключение comments {post_id}: {e}")

    # Shares (/post_id?fields=shares)
    url_shares = f"https://graph.facebook.com/{post_id}"
    params_shares = {
        "fields": "shares",
        "access_token": FACEBOOK_PAGE_ACCESS_TOKEN
    }
    try:
        r_shares = requests.get(url_shares, params=params_shares, proxies=proxy_dict, timeout=30)
        if r_shares.status_code == 200:
            data_sh = r_shares.json()
            shares_obj = data_sh.get("shares", {})
            shares_count = shares_obj.get("count", 0)
        else:
            log_action("get_post_insights", f"Ошибка shares {post_id}: {r_shares.status_code} {r_shares.text}")
    except Exception as e:
        log_action("get_post_insights", f"Исключение shares {post_id}: {e}")

    return likes_count, comments_count, shares_count


def get_post_comments(post_id: str, limit: int = 50):
    """
    Возвращает список (id, message, from_name) последних limit комментариев.
    """
    url = f"https://graph.facebook.com/{post_id}/comments"
    params = {
        "access_token": FACEBOOK_PAGE_ACCESS_TOKEN,
        "fields": "id,message,from",
        "limit": limit
    }
    try:
        r = requests.get(url, params=params, proxies=proxy_dict, timeout=30)
        if r.status_code == 200:
            data = r.json()
            comments = data.get("data", [])
            result = []
            for c in comments:
                cid = c.get("id")
                msg = c.get("message", "")
                author_name = c.get("from", {}).get("name", "NoName")
                result.append((cid, msg, author_name))
            return result
        else:
            log_action("get_post_comments", f"Ошибка {r.status_code} {r.text}")
            return []
    except Exception as e:
        log_action("get_post_comments", f"Исключение: {e}")
        return []


def like_comment(comment_id: str):
    """
    Ставим лайк на комментарий.
    """
    url = f"https://graph.facebook.com/{comment_id}/likes"
    params = {
        "access_token": FACEBOOK_PAGE_ACCESS_TOKEN
    }
    try:
        r = requests.post(url, params=params, proxies=proxy_dict, timeout=30)
        if r.status_code == 200:
            log_action("like_comment", f"Лайк на комментарий {comment_id}")
        else:
            log_action("like_comment", f"Ошибка {comment_id}: {r.status_code} {r.text}")
    except Exception as e:
        log_action("like_comment", f"Исключение {comment_id}: {e}")


def reply_to_comment(comment_id: str, text: str):
    """
    Отвечаем на комментарий по ID.
    """
    url = f"https://graph.facebook.com/{comment_id}/comments"
    params = {
        "access_token": FACEBOOK_PAGE_ACCESS_TOKEN
    }
    data = {
        "message": text
    }
    try:
        r = requests.post(url, params=params, data=data, proxies=proxy_dict, timeout=30)
        if r.status_code == 200:
            log_action("reply_to_comment", f"Ответ на комментарий {comment_id}")
        else:
            log_action("reply_to_comment", f"Ошибка {comment_id}: {r.status_code} {r.text}")
    except Exception as e:
        log_action("reply_to_comment", f"Исключение {comment_id}: {e}")


# =============================================================================
#               ЛОГИКА МОНИТОРИНГА ПОСТОВ (ФОНОВЫЙ ПОТОК)
# =============================================================================

def monitor_posts_loop():
    """
    Фоновый цикл: каждые CHECK_INTERVAL секунд:
      1) Загружаем список наших постов (posted_posts.txt).
      2) Для каждого поста собираем статистику (likes, comments, shares) и пишем в Excel.
      3) Ищем новые комментарии => лайкаем, отвечаем по REPLY_TEMPLATE, записываем в processed_comments.txt
    """
    while True:
        # 1) Список постов
        post_ids = load_posted_posts()

        # 2) Для каждого поста - обновить статистику, найти новые комментарии
        processed_ids = load_processed_comments()  # чтобы не обрабатывать повторно

        for post_id in post_ids:
            # Получаем (likes, comments, shares) -> обновляем Excel
            likes_count, comments_count, shares_count = get_post_insights(post_id)
            update_post_stats(post_id, likes=likes_count, comments=comments_count, shares=shares_count)

            # Получаем последние 50 комментариев
            comment_list = get_post_comments(post_id, limit=50)

            # Смотрим, какие из них новые
            for (cid, ctext, cauthor) in comment_list:
                if cid not in processed_ids:
                    # Ставим лайк
                    like_comment(cid)
                    # Отвечаем по шаблону
                    reply_text = REPLY_TEMPLATE.format(author=cauthor, comment=ctext)
                    reply_to_comment(cid, reply_text)
                    # Добавляем в список обработанных
                    save_processed_comment(cid)

        time.sleep(CHECK_INTERVAL)


# =============================================================================
#           ЛОГИКА ЧТЕНИЯ КОНСОЛИ (/post ТЕКСТ)
# =============================================================================

def console_loop():
    """
    Бесконечный цикл чтения строк из консоли.
    Если строка начинается с "/post ", публикуем пост (не останавливая мониторинг).
    """
    while True:
        line = sys.stdin.readline()
        if not line:
            time.sleep(1)
            continue

        line = line.strip()
        if line.lower().startswith("/post "):
            # извлекаем текст поста
            post_text = line[6:].strip()
            if post_text:
                result = post_to_facebook(post_text)
                if "_" in result or result.isdigit():
                    print(f"Новый пост опубликован: {result}")
                else:
                    print(f"Ошибка публикации: {result}")
            else:
                print("Не указан текст поста. Формат: /post Текст_поста")


# =============================================================================
#                                 MAIN
# =============================================================================

if __name__ == "__main__":
    # Инициализируем Excel-файл статистики (если нет)
    init_stats_xlsx()

    # Запускаем поток мониторинга
    monitor_thread = threading.Thread(target=monitor_posts_loop, daemon=True)
    monitor_thread.start()

    # Запускаем цикл чтения консоли в основном потоке
    console_loop()























